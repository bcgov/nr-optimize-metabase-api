# -------------------------------------------------------------------------------
# Name:        push_h_sfp_to_metabase_v2.py
# Purpose:     the purpose of the script is to ETL H drive and Group Share data into metabase, as well as produce
#              Divisional Storage Reports based on the H Drive information for distribution to Ministry Executives.
#              1.) Read in xlsx data
#              2.) Transform the data
#              3.) Insert H Drive and Group Share data to Metabase
#              4.) Output formatted Divisional Storage Reports as Excel file(s) to local machine
#
# Author:      PPLATTEN, HHAY
#
# Created:     2022
# Copyright:   (c) Optimization Team 2022
# Licence:     mine
#
#
# usage: push_h_sfp_to_metabase_v2.py
# requirements:
#              1.) Must have the H Drive / Group Share xlsx file(s) in same folder as script.
#              2.) Must have an open port to metabase database.
# -------------------------------------------------------------------------------

import glob
import ldap_helper as ldap
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, DEFAULT_FONT
from openpyxl.formatting.rule import FormulaRule
import pandas as pd
import os
import sys
import psycopg2
import push_postgres_constants as constants
from division_renames import (
    all_division_renames,
    for_division_acronyms,
    wlrs_division_acronyms,
    env_division_acronyms,
    emli_division_acronyms,
    af_division_acronyms,
    irr_division_acronyms,
)
from dsr_functions import assign_div_acronyms

errors = {
    "department": [],
    "physicalDeliveryOfficeName": [],
    "other": [],
    "not_found": [],
}

ministry_renames = {
    "AGRI": "AF",
    "AFF": "AF",
    "FLNR": "FOR",
    "EMPR": "EMLI",
    "MEM": "EMLI",
    "ABR": "IRR",
    "LWRS": "WLRS",
}

delete_before_insert = False


def get_ad_attribute(idir, ldap_util, conn, attribute="givenName"):
    attributes = [attribute]
    ad_info = None

    try:
        # Connect to AD to get user info
        ad_info = ldap_util.getADInfo(idir, conn, attributes)
    except (Exception, AttributeError):
        errors[attribute].append(idir)
        return None

    if ad_info is None or attribute not in ad_info:
        # attribute was not found, if it's an attribute that we track, append it.
        if attribute in errors:
            errors[attribute].append(idir)
        return None
    else:
        return ad_info[attribute]


def manipulate_frame(frame, ministryname, datestamp):
    # remove rows where User ID is blank
    frame = frame.dropna(thresh=1)

    # add ministry and date columns
    frame = frame.assign(ministry=ministryname)
    frame = frame.assign(date=datestamp)

    # update ministry acronyms
    for original_name in ministry_renames:
        frame["ministry"] = frame["ministry"].apply(
            lambda x: x.replace(original_name, ministry_renames[original_name])
        )

    # remove the header row -- assume it's the first
    frame = frame[1:]

    return frame


def add_column(frame, column_ad_name, column_postgress_name):
    ldap_util = ldap.LDAPUtil()
    conn = ldap_util.getLdapConnection()

    values = []
    for i in range(len(frame.values)):
        idir = frame.values[i][0]
        if type(idir) == str and idir != "User ID":
            values.append(get_ad_attribute(idir, ldap_util, conn, column_ad_name))
        else:
            values.append("")
    frame[column_postgress_name] = values

    return frame


# gets all .xlsx files from the python files directory
def get_records_from_xlsx(sheet_name):

    # grab all the .xlsx file names in the python file's directory
    current_file_path = os.path.dirname(os.path.realpath(__file__))
    excel_names = glob.glob(os.path.join(current_file_path, "*.xlsx"))

    frames = []

    # pull the data out of each xlsx, and aggregate it
    for name in excel_names:

        # find the ministry based on the filename
        ministries = [
            "agri",
            "empr",
            "env",
            "irr",
            "flnr",
            "emli",
            "aff",
            "mem",
            "abr",
            "fpro",
            "bcws",
            "af",
            "for",
            "lwrs",
            "wlrs",
        ]
        for ministry_acronym in ministries:
            if ministry_acronym in name.lower():
                ministryname = ministry_acronym.upper()
                break

        # find the date based on the filename
        pos = name.rfind("\\") + 1
        datestamp = name[pos : pos + 7].strip() + "-01"

        # read in xlsx, turn into dataframes
        excelsheet = pd.ExcelFile(name)

        # get and touch up data for all sheets which include the selected sheet name
        for current_sheet_name in excelsheet.sheet_names:
            if sheet_name.lower() in current_sheet_name.lower():
                print(f"Working on file {name} sheet {current_sheet_name}")
                frame = excelsheet.parse(
                    current_sheet_name, header=None, index_col=None
                )
                frame = manipulate_frame(frame, ministryname, datestamp)

                frames.append(frame)

    # Merge the datasets together
    combined = pd.concat(frames)

    # Log the users not found in LDAP
    with open("not_found.txt", "w") as f:
        for idir in errors["not_found"]:
            f.write(f"{idir}\n")

    # add headers back in
    if sheet_name.lower() == "home drives":
        combined.columns = [
            "idir",
            "displayname",
            "mailboxcode",
            "fname",
            "lname",
            "division",
            "branch",
            "email",
            "datausage",
            "ministry",
            "date",
        ]

        # also fill blank displaynames with idir
        combined.displayname.fillna(combined.idir, inplace=True)
    elif sheet_name.lower() == "group shares":
        combined.columns = ["sharename", "server", "datausage", "ministry", "date"]

    # flatten data into tuples
    record_tuples = []
    for row in combined.values:
        if type(row[0]) == str:
            record_tuples.append(tuple(row))

    return record_tuples


def get_conn():
    return psycopg2.connect(
        host="localhost",
        database=constants.POSTGRES_DB_NAME,
        user=constants.POSTGRES_USER,
        password=constants.POSTGRES_PASS,
    )


# inserts h drive tuples into metabase
def insert_h_drive_records_to_metabase(record_tuples):

    insert_tuples = []
    for tup in record_tuples:
        # idir, displayname, datausage, division, branch, ministry, date
        new_tup = (tup[0], tup[1], tup[8], tup[5], tup[6], tup[9], tup[10])
        insert_tuples.append(new_tup)

    conn = None
    try:
        # Open a connection
        conn = get_conn()
        cur = conn.cursor()

        if delete_before_insert:
            print("Deleting old h drive data")
            cur.execute("DELETE FROM hdriveusage")
            print("Delete from hdriveusage complete")

        print("Inserting new h drive data")
        cur.executemany(
            """
            INSERT INTO hdriveusage (idir, displayname, datausage, division, branch, ministry, date)
            VALUES (%s, %s, %s, %s, %s, %s, %s);
        """,
            insert_tuples,
        )
        print("Insert to hdriveusage Complete")

        # close the communication with the PostgreSQL
        conn.commit()
        cur.close()
    except (Exception, psycopg2.DatabaseError) as error:
        print(error)
    finally:
        if conn is not None:
            conn.close()
            print("Database connection closed.")


# inserts group share tuples into metabase
def insert_group_share_records_to_metabase(record_tuples):
    conn = None
    try:
        # Open a connection
        conn = get_conn()
        cur = conn.cursor()

        if delete_before_insert:
            print("Deleting old group share data")
            cur.execute("DELETE FROM sfpmonthly")
            print("Delete from sfpmonthly complete")

        print("Inserting new group share data")
        cur.executemany(
            """
            INSERT INTO sfpmonthly (sharename, server, datausage, ministry, date)
            VALUES (%s, %s, %s, %s, %s);
        """,
            record_tuples,
        )
        print("Insert to sfpmonthly Complete")

        # close the communication with the PostgreSQL
        conn.commit()
        cur.close()
    except (Exception, psycopg2.DatabaseError) as error:
        print(error)
    finally:
        if conn is not None:
            conn.close()
            print("Database connection closed.")


def create_ministry_reports_simple(record_tuples):
    # Arrange tuples by ministry/workbook
    ministry_dict = {}
    for tup in record_tuples:
        tup = tup
        if tup[0].lower() == "soft deleted home drives":
            continue
        ministry = tup[9] if isinstance(tup[9], str) else "None"
        if ministry not in ministry_dict:
            ministry_dict[ministry] = []
        ministry_dict[ministry].append(tup)

    # Create an excel workbook for each ministry
    for ministry in ministry_dict:
        # Format for excel output
        tups = ministry_dict[ministry]
        df_array = []
        for tup in tups:
            # Yes if consumption is over 1.5gb, otherwise user is not over the limit
            over_limit = "Y" if tup[8] > 1.5 else "N"
            # Convert to array for use in data frame, column names below
            df_array.append([tup[9], tup[5], tup[6], f"{tup[4]}, {tup[3]}", over_limit])

        # apply division renames to correct bad data in AD
        for row in df_array:
            if row[1] in all_division_renames:
                row[1] = all_division_renames[row[1]]

        # Convert to dataframe and add column names
        df1 = pd.DataFrame(
            df_array,
            columns=[
                "Ministry",
                "Division",
                "Branch",
                "Display Name",
                "Over Limit (1.5gb)",
            ],
        )

        assign_div_acronyms(df1, "AF", af_division_acronyms)
        assign_div_acronyms(df1, "EMLI", emli_division_acronyms)
        assign_div_acronyms(df1, "ENV", env_division_acronyms)
        assign_div_acronyms(df1, "FOR", for_division_acronyms)
        assign_div_acronyms(df1, "IRR", irr_division_acronyms)
        assign_div_acronyms(df1, "WLRS", wlrs_division_acronyms)
        df1.drop("Ministry", axis=1, inplace=True)
        df1.sort_values(
            ["Division", "Branch", "Over Limit (1.5gb)", "Display Name"], inplace=True
        )
        # group dataframes by Division
        for div, group in df1.groupby(by=["Div_Acronym"]):

            # Set file name
            ministry_upper = ministry.upper()
            div_acronym = group["Div_Acronym"].values[0]
            # div_acronym = get_div_acronym(div_name, )
            yyyy_mm_dd = tups[0][10]
            tups = tups
            file_name = f"{ministry_upper}_{div_acronym}_DSR_{yyyy_mm_dd}.xlsx"
            # Convert dataframe to formatted Excel file
            wb = Workbook()
            ws = wb.active
            ws.title = div_acronym
            DEFAULT_FONT.name = "Calibri"
            _font = Font(name="Calibri", sz=11)
            {k: setattr(DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}
            for r in dataframe_to_rows(group, index=False, header=True):
                ws.append(r)
            # Expand the columns
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max(
                            (dims.get(cell.column_letter, 0), len(str(cell.value)))
                        )
            for col, value in dims.items():
                ws.column_dimensions[col].width = value
            # Bold the header row
            for cell in ws["1:1"]:
                cell.font = Font(bold=True)
            # Highlight any cells where the data usage is over the specified limit
            for cell in ws["A"] + ws[1]:
                redFill = PatternFill(
                    start_color="FFEE1111", end_color="FFEE1111", fill_type="solid"
                )
                ws.conditional_formatting.add(
                    "D2:D8000",
                    FormulaRule(
                        formula=['NOT(ISERROR(SEARCH("Y",D2)))'],
                        stopIfTrue=True,
                        fill=redFill,
                    ),
                )

            # Add the disclaimer and a blank row for spacing
            ws.insert_rows(0)
            ws.insert_rows(0)
            ws.insert_rows(0)
            disclaimer_line_1 = "The content of this Division Storage Report is confidential and intended for the recipient specified. You should only see data for your work area. If you received information about employees outside your work area, please delete that data immediately. Thank you for your cooperation and understanding."  # noqa: E501
            ws.append({1: disclaimer_line_1})
            disclaimer_line_2 = "Our Privacy Impact Assessment requires us to alert users of this report of potential inaccuracies. If you identify any discrepancies between the report and an employee’s actual H Drive, please contact the Optimization Team at NRIDS.Optimize@gov.bc.ca."  # noqa: E501
            ws.append({1: disclaimer_line_2})
            disclaimer_cell_range = "A"+str(ws.max_row-1)+":A"+str(ws.max_row)
            ws.move_range(disclaimer_cell_range, rows=-(ws.max_row-2))
            for cell in ws["1:1"]:
                cell.font = Font(bold=True)
            for cell in ws["2:2"]:
                cell.font = Font(bold=True)

            # save the workbook - makes directory if it doesn't already exist
            path = f"C:/git/output/{ministry_upper}"
            if not os.path.exists(path):
                os.makedirs(path)
            wb.save(f"{path}/{file_name}")


if __name__ == "__main__":
    if "-d" in sys.argv:
        delete_before_insert = True
    record_tuples = get_records_from_xlsx("home drives")
    create_ministry_reports_simple(record_tuples)
    insert_h_drive_records_to_metabase(record_tuples)

    record_tuples = get_records_from_xlsx("group shares")
    insert_group_share_records_to_metabase(record_tuples)
