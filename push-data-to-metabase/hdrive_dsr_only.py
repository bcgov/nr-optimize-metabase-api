# import python libraries
import os
import sys
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, DEFAULT_FONT
from openpyxl.formatting.rule import FormulaRule
from division_renames import (
    all_division_renames,
    for_division_acronyms,
    wlrs_division_acronyms,
    env_division_acronyms,
    emli_division_acronyms,
    af_division_acronyms,
    irr_division_acronyms,
)
from dsr_functions import assign_div_acronyms

ministry_renames = {
    "AGRI": "AF",
    "AFF": "AF",
    "ALC": "AF",
    "FLNR": "FOR",
    "BCWS": "FOR",
    "EMPR": "EMLI",
    "MEM": "EMLI",
    "ABR": "IRR",
    "LWRS": "WLRS",
    "EAO": "ENV",
}


def manipulate_frame(frame, ministryname, datestamp):
    # remove rows where User ID is blank
    frame = frame.dropna(thresh=1)

    # add ministry and date columns
    frame = frame.assign(ministry=ministryname)
    frame = frame.assign(date=datestamp)

    # update ministry acronyms
    for original_name in ministry_renames:
        if original_name not in ministry_renames:
            frame["ministry"] = frame["ministry"].apply(
                lambda x: x.replace(original_name, ministry_renames["NON STANDARD"])
            )

    # remove the header row -- assume it's the first
    frame = frame[1:]

    return frame


# gets all .xlsx files from the python files directory
def get_records_from_xlsx(sheet_name):
    # grab all the .xlsx file names in the python file's directory
    current_file_path = os.path.dirname(os.path.realpath(__file__))
    excel_names = glob.glob(os.path.join(current_file_path, "*.xlsx"))

    frames = []

    # pull the data out of each xlsx, and aggregate it
    for name in excel_names:
        # find the ministry based on the filename
        ministries = [
            "env",
            "irr",
            "emli",
            "af",
            "for",
            "wlrs",
        ]
        for ministry_acronym in ministries:
            if ministry_acronym in name.lower():
                ministryname = ministry_acronym.upper()
                break

        # find the date based on the filename
        pos = name.rfind("\\") + 1
        datestamp = name[pos : pos + 7].strip() + "-15"

        # find the ministry based on the filename so it can be assigned to soft-deleted home drives
        sheet_min = name[12:5]

        # read in xlsx, turn into dataframes
        excelsheet = pd.ExcelFile(name)

        # get and touch up data for all sheets which include the selected sheet name
        for current_sheet_name in excelsheet.sheet_names:
            if sheet_name.lower() in current_sheet_name.lower():
                print(f"Working on file {name} sheet {current_sheet_name}")
                frame = excelsheet.parse(
                    current_sheet_name, header=None, index_col=None
                )
                frame = manipulate_frame(frame, ministryname, datestamp)

                frames.append(frame)

    # Merge the datasets together
    combined = pd.concat(frames)

    # add headers back in
    if sheet_name.lower() == "home drives":
        combined.columns = [
            "idir",
            "displayname",
            "mailboxcode",
            "fname",
            "lname",
            "division",
            "branch",
            "email",
            "datausage",
            "ministry",
            "date",
        ]

        # fill blank displaynames with idir
        combined.displayname.fillna(combined.idir, inplace=True)

        combined.loc[
            combined["displayname"] == "NO IDIR USER RECORD FOUND", "branch"
        ] = "Not Found"

        # fill certain divisions with "Not Found"
        combined.loc[
            combined["displayname"] == "NO IDIR USER RECORD FOUND", "division"
        ] = "Not Found"

        # handle EAO extensions that have no division
        combined.loc[combined["displayname"].str.contains("EAO:EX"), "division"] = (
            "Environmental Assessment Office"
        )

        # set Div & Branch to blank for Soft deleted Home Drives
        combined.loc[combined["idir"] == "Soft deleted Home Drives", "division"] = ""
        combined.loc[combined["idir"] == "Soft deleted Home Drives", "branch"] = ""

        # check mailbox org code for non-NRM users, leaving CSNR as-is for now
        combined.loc[combined["mailboxcode"] == "RBCM", "ministry"] = "NON STANDARD"
        combined.loc[combined["mailboxcode"] == "CITZ", "ministry"] = "NON STANDARD"
        combined.loc[combined["mailboxcode"] == "TACS", "ministry"] = "NON STANDARD"
        combined.loc[combined["mailboxcode"] == "FIN", "ministry"] = "NON STANDARD"
        combined.loc[combined["mailboxcode"] == "ECC", "ministry"] = "NON STANDARD"

        # check mailbox org code for NRM assignment
        combined.loc[combined["mailboxcode"] == "AF", "ministry"] = "AF"
        combined.loc[combined["mailboxcode"] == "EMLI", "ministry"] = "EMLI"
        combined.loc[combined["mailboxcode"] == "ENV", "ministry"] = "ENV"
        combined.loc[combined["mailboxcode"] == "FOR", "ministry"] = "FOR"
        combined.loc[combined["mailboxcode"] == "BCWS", "ministry"] = "FOR"
        combined.loc[combined["mailboxcode"] == "IRR", "ministry"] = "IRR"
        combined.loc[combined["mailboxcode"] == "DAS", "ministry"] = "IRR"
        combined.loc[combined["mailboxcode"] == "WLRS", "ministry"] = "WLRS"
        combined.loc[combined["mailboxcode"] == "", "ministry"] = sheet_min

        # add limit column
        combined["Over Limit (1.5gb)"] = ""
        combined.loc[combined["datausage"] <= 1.5, "Over Limit (1.5gb)"] = "N"
        combined.loc[combined["datausage"] > 1.5, "Over Limit (1.5gb)"] = "Y"

        # flatten data into tuples
    record_tuples = []
    for row in combined.values:
        if type(row[0]) == str:
            record_tuples.append(tuple(row))

    return record_tuples


def create_ministry_reports_simple(record_tuples):
    # Arrange tuples by ministry/workbook
    ministry_dict = {}
    for tup in record_tuples:
        tup = tup
        if tup[0].lower() == "soft deleted home drives":
            continue
        ministry = tup[9] if isinstance(tup[9], str) else "None"
        if ministry not in ministry_dict:
            ministry_dict[ministry] = []
        ministry_dict[ministry].append(tup)

    # Create an excel workbook for each ministry
    for ministry in ministry_dict:
        # Format for excel output
        tups = ministry_dict[ministry]
        df_array = []
        for tup in tups:
            # Convert to array for use in data frame, column names below
            df_array.append([tup[9], tup[5], tup[6], tup[1], tup[11]])

        # apply division renames to correct bad data in AD
        for row in df_array:
            if row[1] in all_division_renames:
                row[1] = all_division_renames[row[1]]

        # Convert to dataframe and add column names
        df1 = pd.DataFrame(
            df_array,
            columns=[
                "Ministry",
                "Division",
                "Branch",
                "Display Name",
                "Over Limit (1.5gb)",
            ],
        )

        assign_div_acronyms(df1, "AF", af_division_acronyms)
        assign_div_acronyms(df1, "EMLI", emli_division_acronyms)
        assign_div_acronyms(df1, "ENV", env_division_acronyms)
        assign_div_acronyms(df1, "FOR", for_division_acronyms)
        assign_div_acronyms(df1, "IRR", irr_division_acronyms)
        assign_div_acronyms(df1, "WLRS", wlrs_division_acronyms)

        df1.drop("Ministry", axis=1, inplace=True)
        # drop rows where there is no display name (denoting a resource account)
        df1.drop(df1[df1["Display Name"] == "nan, nan"].index, inplace=True)

        df1.sort_values(
            ["Division", "Branch", "Over Limit (1.5gb)", "Display Name"], inplace=True
        )

        # group dataframes by Division
        for div, group in df1.groupby(by=["Div_Acronym"]):
            # Set file name
            ministry_upper = ministry.upper()
            div_acronym = group["Div_Acronym"].values[0]
            # div_acronym = get_div_acronym(div_name, )
            yyyy_mm_dd = tups[0][10]
            tups = tups
            file_name = f"{ministry_upper}_{div_acronym}_DSR_{yyyy_mm_dd}.xlsx"
            # Convert dataframe to formatted Excel file
            wb = Workbook()
            ws = wb.active
            ws.title = div_acronym
            DEFAULT_FONT.name = "Calibri"
            _font = Font(name="Calibri", sz=11)
            {k: setattr(DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}
            for r in dataframe_to_rows(group, index=False, header=True):
                ws.append(r)
            # Expand the columns
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max(
                            (dims.get(cell.column_letter, 0), len(str(cell.value)))
                        )
            for col, value in dims.items():
                ws.column_dimensions[col].width = value
            # Bold the header row
            for cell in ws["1:1"]:
                cell.font = Font(bold=True)
            # Highlight any cells where the data usage is over the specified limit
            for cell in ws["A"] + ws[1]:
                redFill = PatternFill(
                    start_color="FFEE1111", end_color="FFEE1111", fill_type="solid"
                )
                ws.conditional_formatting.add(
                    "D2:D8000",
                    FormulaRule(
                        formula=['NOT(ISERROR(SEARCH("Y",D2)))'],
                        stopIfTrue=True,
                        fill=redFill,
                    ),
                )

            # Add the disclaimer and a blank row for spacing
            def add_disclaimer_line_to_top(text):
                ws.insert_rows(0)
                ws.append({1: text})
                disclaimer_cell_range = "A" + str(ws.max_row) + ":A" + str(ws.max_row)
                ws.move_range(disclaimer_cell_range, rows=-(ws.max_row - 1))
                for cell in ws["1:1"]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(
                        start_color="003366", end_color="003366", fill_type="solid"
                    )

            disclaimer_line_1 = "The content of this Division Storage Report is confidential and intended for the recipient specified. You should only see data for your work area."  # noqa: E501
            disclaimer_line_2 = "If you received information about employees outside your work area, please delete that data immediately. Thank you for your cooperation and understanding."  # noqa: E501
            disclaimer_line_3 = "Our Privacy Impact Assessment requires us to alert users of this report of potential inaccuracies."  # noqa: E501
            disclaimer_line_4 = "If you identify any discrepancies between the report and an employee’s actual H Drive, please contact the Optimization Team at NRIDS.Optimize@gov.bc.ca."  # noqa: E501
            ws.insert_rows(0)
            add_disclaimer_line_to_top(disclaimer_line_4)
            add_disclaimer_line_to_top(disclaimer_line_3)
            add_disclaimer_line_to_top(disclaimer_line_2)
            add_disclaimer_line_to_top(disclaimer_line_1)
            # merge cells from A1 to E4
            ws.merge_cells("A1:E1")
            ws.merge_cells("A2:E2")
            ws.merge_cells("A3:E3")
            ws.merge_cells("A4:E4")

            # save the workbook - makes directory if it doesn't already exist
            path = f"C:/Git_Repo/Output/{ministry_upper}"
            if not os.path.exists(path):
                os.makedirs(path)
            wb.save(f"{path}/{file_name}")


record_tuples = get_records_from_xlsx("home drives")
create_ministry_reports_simple(record_tuples)
