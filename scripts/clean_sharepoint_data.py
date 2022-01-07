# -------------------------------------------------------------------------------
# Name:        clean_sharepoint_data.py
# Purpose:     the purpose of the script is to clean the SharePoint report data:
#              1.) Add a date column
#              2.) Extract URL from site name to a separate column
#              3.) Calculate the data usage in GB
#
# Author:      HHAY, JMONTEBE, PPLATTEN
#
# Created:     2022
# Copyright:   (c) Optimization Team 2022
# Licence:     mine
#
#
# usage: clean_sharepoint_data.py -i <input_directory> -f <file_name> -d <report_date>
# example:  clean_sharepoint_data.py -i "J:\Scripts\Python\data_to_process" -f "2021-12-13_CollectionStorageMetrics.xlsx" -d 2021-12-01
# -------------------------------------------------------------------------------

import sys
import os
import argparse
import pandas as pd
from openpyxl import load_workbook


def main(argv):
    # take a single input directory argument in
    # take a single file name argument in
    # take a single date argument in
    # print ('Number of arguments:', len(sys.argv), 'arguments.')
    # print ('Argument List:', str(sys.argv))

    inputdirectory = ""
    filename = ""
    date = ""
    syntaxcmd = "Insufficient number of commands passed: clean_sharepoint_data.py -i <input_dir> -f <file_name> -d <datecol>"

    if len(sys.argv) < 3:
        print(syntaxcmd)
        sys.exit(1)

    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-i",
        "--input",
        dest="inputdirectory",
        required=True,
        help="path to directory containing SharePoint usage reports",
        metavar="string",
        type=str,
    )
    parser.add_argument(
        "-f",
        "--file",
        dest="filename",
        required=True,
        help="file name of the SharePoint usage report",
        metavar="string",
        type=str,
    )
    parser.add_argument(
        "-d",
        "--date",
        dest="date",
        required=True,
        help="YYYY-MM-DD that the SharePoint report was created by Web Analyst, for the date column. Always use 01 as the day",
        metavar="string",
        type=str,
    )
    args = parser.parse_args()

    inputdirectory = args.inputdirectory
    filename = args.filename
    date = args.date

    # provide full path to excel file
    pathname = os.path.join(inputdirectory, filename)
    # load excel file
    workbook = pd.ExcelFile(pathname)

    # join all excel worksheets in file as dataframe, specifying desired columns
    df = pd.concat(
        pd.read_excel(
            workbook,
            sheet_name=None,
            index_col=None,
            usecols="B:C,F",
        ),
        ignore_index=True,
    )

    # rename dataframe columns
    df2 = df.rename(
        {
            "Name": "sitename",
            "Total Size": "datausage",
            "Last Modified": "lastmodified",
        },
        axis=1,
    )

    # manipulate dataframe columns
    df2["date"] = date
    df2["datausage"] = df2["datausage"].apply(
        lambda x: x.replace("Not Available", "0000")
    )
    df2["unit"] = df2.datausage.str[-2:]
    df2["datausage"] = df2["datausage"].str[:-3]
    df2["sitename"] = [str(x).replace("–", "-") for x in df2["sitename"]]
    df2["sitename"] = [str(x).replace("í", "i") for x in df2["sitename"]]
    df2["sitename"] = [str(x).replace("á", "a") for x in df2["sitename"]]
    df2["sitename"] = [str(x).replace("’", "'") for x in df2["sitename"]]

    # send dataframe columns to list for further manipulation
    datausage_list = df2["datausage"].tolist()
    lastmod_list = df2["lastmodified"].tolist()
    sitename_list = df2["sitename"].tolist()

    # using openpyxl to pull hyperlink values that pandas cannot handle (also sheet titles for convenience)
    excel_file = load_workbook(filename=(os.path.join(inputdirectory, filename)))

    url = []
    collection = []

    for sheet in excel_file.worksheets:
        # iterate through excel sheets
        # pull out SharePoint URL from the site name column, collection name from sheet name
        for i in range(2, sheet.max_row + 1):
            try:
                url.append(sheet.cell(row=i, column=2).hyperlink.target)
                collection.append(sheet.title)
            except:
                url.append(sheet.cell(row=i, column=2).value)

    # remove blank rows from the column data extracted using openpyxl
    url_list = []
    for val in url:
        if val is not None:
            url_list.append(val)

    collection_list = []
    for val in collection:
        if val is not None:
            collection_list.append(val)

    # assign columns to specific dataframe type for altering
    df3 = pd.DataFrame(url_list)
    df4 = pd.DataFrame(collection_list)
    df5 = pd.to_numeric(datausage_list)
    df6 = pd.DataFrame(lastmod_list)
    df7 = pd.DataFrame(sitename_list)

    nan_value = float("NaN")
    # Convert NaN values to empty string
    df3.replace("", nan_value, inplace=True)
    df4.replace("", nan_value, inplace=True)
    df6.replace("", nan_value, inplace=True)
    df7.replace("", nan_value, inplace=True)

    # drop blank rows
    df3.dropna(inplace=True)
    df4.dropna(inplace=True)
    df6.dropna(inplace=True)
    df7.dropna(inplace=True)

    # assign the altered columns back to df2
    df2["url"] = df3
    df2["collection"] = df4
    df2["datausage"] = df5
    df2["lastmodified"] = df6
    df2["sitename"] = df7

    # send all the new and altered columns to a final dataframe
    df_final = df2

    # convert all data sizes to GB
    df_final.loc[df_final["unit"] == "MB", "datausage"] = df_final["datausage"] / 1000
    df_final.loc[df_final["unit"] == "KB", "datausage"] = (
        df_final["datausage"] / 1000000
    )

    # calculate the cost column
    df_final["datacost"] = df_final["datausage"] * 60

    # re-order the dataframe columns
    df_final = df_final[
        [
            "collection",
            "sitename",
            "url",
            "datausage",
            "datacost",
            "lastmodified",
            "date",
        ]
    ]

    # send dataframe to csv output
    output_file = date + "_SharePoint_Data.csv"
    df_final.to_csv(output_file, encoding="utf-8", index=False)


if __name__ == "__main__":
    main(sys.argv[1:])
